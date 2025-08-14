from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
import psycopg2
from datetime import datetime, timedelta
import pytz
import os
from flask import send_file
from psycopg2.extras import DictCursor
import cloudinary
import cloudinary.uploader

cloudinary.config(
    cloud_name='dcbdjnpzo',
    api_key='381622333637456',
    api_secret='P1Pzvu85aCR02HuRSCnz76yzrgg'
)

app = Flask(__name__)
#app.secret_key = 'tu_clave_secreta_aqui'  # Necesario para usar local
import os
app.secret_key = os.getenv("SECRET_KEY", "change-me")


# Configuración de la conexión a PostgreSQL
from psycopg2.extras import DictCursor

def get_db_connection():
    dsn = os.getenv("DATABASE_URL")
    if not dsn:
        raise RuntimeError("Falta DATABASE_URL")
    return psycopg2.connect(dsn, cursor_factory=DictCursor, sslmode="require")



# Crear tabla de usuarios si no existe
def crear_tabla_usuarios():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user'
        )
    ''')
    conn.commit()
    conn.close()

# Llamar a la función para crear la tabla de usuarios al iniciar la aplicación
crear_tabla_usuarios()

# Función para crear la tabla `equipos` si no existe
def crear_tabla_equipos():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS equipos_sj (
            id SERIAL PRIMARY KEY,
            tipo_reparacion TEXT NOT NULL,
            marca TEXT NOT NULL,
            modelo TEXT NOT NULL,
            tecnico TEXT NOT NULL,
            monto REAL NOT NULL,
            nombre_cliente TEXT NOT NULL,
            telefono TEXT NOT NULL,
            nro_orden TEXT NOT NULL,
            fecha TEXT NOT NULL,
            hora TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

# Llamar a la función para crear la tabla al iniciar la aplicación
crear_tabla_equipos()

# Proteger rutas que requieren autenticación
def login_required(f):
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            flash('Debes iniciar sesión para acceder a esta página.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Ruta principal (redirige al login si no está autenticado)
@app.route('/')
def index():
    if 'username' in session:
        return redirect(url_for('inicio'))  # Redirige a la página principal del sistema
    return redirect(url_for('login'))  # Redirige al login si no está autenticado


# Ruta para el login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'username' in session:
        return redirect(url_for('inicio'))  # Redirige a la página principal si ya está autenticado

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM usuarios WHERE username = %s', (username,))
        user = cursor.fetchone()
        conn.close()

        if user and user['password'] == password:
            session['username'] = user['username']
            session['role'] = user['role']

            return redirect(url_for('inicio'))  # Redirige a la página principal después del login
        else:
            flash('Usuario o contraseña incorrectos', 'error')

    return render_template('login.html')

# Ruta para la página principal del sistema (después del login)
@app.route('/inicio')
def inicio():
    if 'username' not in session:
        return redirect(url_for('login'))  # Redirige al login si no está autenticado
    return render_template('inicio.html')

# Ruta para el logout
@app.route('/logout')
def logout():
    session.pop('username', None)
    session.pop('role', None)
    flash('Has cerrado sesión correctamente.', 'success')
    return redirect(url_for('login'))

# Ruta para registrar ventas
@app.route('/registrar_venta', methods=['GET', 'POST'])
def registrar_venta():
    conn = get_db_connection()
    cursor = conn.cursor()

    if 'carrito' not in session:
        session['carrito'] = []

    if request.method == 'POST':
        if 'buscar' in request.form:
            busqueda = request.form['busqueda']
            cursor.execute('''
                SELECT id, nombre, codigo_barras, num, stock, precio, precio_revendedor FROM productos_sj
                WHERE codigo_barras = %s OR nombre ILIKE %s OR num ILIKE %s
            ''', (busqueda, f'%{busqueda}%', f'%{busqueda}%'))
            productos = cursor.fetchall()
            conn.close()
            return render_template('registrar_venta.html', productos=productos, carrito=session['carrito'])

        elif 'agregar' in request.form:
            producto_id = request.form['producto_id']
            cantidad = int(request.form['cantidad'])

            tipo_precio = 'venta'  # por defecto

            cursor.execute('SELECT id, nombre, stock, precio, precio_revendedor FROM productos_sj WHERE id = %s', (producto_id,))
            producto = cursor.fetchone()

            if producto:
                stock = producto['stock']
                if stock >= cantidad:
                    precio_usado = float(producto['precio'])

                    item = {
                        'id': producto['id'],
                        'nombre': producto['nombre'],
                        'precio': precio_usado,
                        'cantidad': cantidad,
                        'tipo_precio': tipo_precio
                    }

                    session['carrito'].append(item)
                    session.modified = True
                else:
                    flash(f'Sin stock suficiente para "{producto["nombre"]}"', 'error')
            else:
                flash('Producto no encontrado', 'error')

        elif 'agregar_manual' in request.form:
            nombre = request.form['nombre_manual']
            precio = float(request.form['precio_manual'])
            cantidad = int(request.form['cantidad_manual'])

            item = {
                'id': None,
                'nombre': nombre,
                'precio': precio,
                'cantidad': cantidad,
                'tipo_precio': 'manual'
            }
            session['carrito'].append(item)
            session.modified = True
            flash(f'Servicio técnico "{nombre}" agregado al carrito', 'success')

        elif 'registrar' in request.form:
            if not session['carrito']:
                flash('Carrito vacío. Agregá productos antes de registrar la venta', 'error')
                return redirect(url_for('registrar_venta'))

            tipo_precio_global = request.form.get('tipo_precio', 'venta')
            tipo_pago = request.form['tipo_pago']
            dni_cliente = request.form['dni_cliente']
            argentina_tz = pytz.timezone('America/Argentina/Buenos_Aires')
            fecha_actual = datetime.now(argentina_tz).strftime('%Y-%m-%d %H:%M:%S')

            for item in session['carrito']:
                producto_id = item['id']
                nombre = item['nombre']
                cantidad = int(item['cantidad'])

                if producto_id is not None:
                    cursor.execute('SELECT precio, precio_revendedor, stock FROM productos_sj WHERE id = %s', (producto_id,))
                    producto = cursor.fetchone()

                    if producto and producto['stock'] >= cantidad:
                        if tipo_precio_global == 'revendedor' and producto['precio_revendedor']:
                            precio = float(producto['precio_revendedor'])
                            tipo_precio = 'revendedor'
                        else:
                            precio = float(producto['precio'])
                            tipo_precio = 'venta'

                        cursor.execute('''
                            INSERT INTO ventas_sj (
                                producto_id, cantidad, fecha, nombre_manual, tipo_pago, dni_cliente, tipo_precio
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (producto_id, cantidad, fecha_actual, nombre, tipo_pago, dni_cliente, tipo_precio))

                        cursor.execute('UPDATE productos_sj SET stock = stock - %s WHERE id = %s', (cantidad, producto_id))
                    else:
                        conn.close()
                        flash(f'Sin stock suficiente para el producto: {nombre}', 'error')
                        return redirect(url_for('registrar_venta'))

                else:
                    # Servicio técnico (manual)
                    precio = float(item['precio'])
                    cursor.execute('''
                        INSERT INTO reparaciones_sj (nombre_servicio, precio, cantidad, tipo_pago, dni_cliente, fecha)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    ''', (nombre, precio, cantidad, tipo_pago, dni_cliente, fecha_actual))

            conn.commit()
            conn.close()
            session.pop('carrito', None)
            flash('Venta registrada con éxito', 'success')
            return redirect(url_for('registrar_venta'))

        elif 'vaciar' in request.form:
            session.pop('carrito', None)
            flash('Carrito vaciado con éxito', 'success')
            return redirect(url_for('registrar_venta'))

    total = sum(float(item['precio']) * int(item['cantidad']) for item in session['carrito'] if 'precio' in item)
    conn.close()
    return render_template('registrar_venta.html', productos=None, carrito=session['carrito'], total=total)








@app.route('/api/carrito/precios_actualizados')
def precios_actualizados():
    tipo_precio = request.args.get('tipo_precio', 'venta')
    carrito = session.get('carrito', [])
    conn = get_db_connection()
    cursor = conn.cursor()

    nuevos_items = []
    for item in carrito:
        if item['id']:
            cursor.execute("SELECT precio, precio_revendedor FROM productos_sj WHERE id = %s", (item['id'],))
            datos = cursor.fetchone()
            precio = float(datos['precio'] if tipo_precio == 'venta' else datos['precio_revendedor'])
            nuevos_items.append({
                'nombre': item['nombre'],
                'cantidad': item['cantidad'],
                'precio': precio
            })
        else:
            nuevos_items.append(item)

    conn.close()
    return jsonify(nuevos_items)




# Ruta para mostrar los productos más vendidos
@app.route('/productos_mas_vendidos')
def productos_mas_vendidos():
    # Conectar a la base de datos
    conn = get_db_connection()
    cursor = conn.cursor()

    # Consulta para obtener los 5 productos más vendidos
    cursor.execute('''
        SELECT nombre, precio, cantidad_vendida 
        FROM productos_sj
        ORDER BY cantidad_vendida DESC 
        LIMIT 5
    ''')
    productos = cursor.fetchall()

    # Calcular el total de ventas
    cursor.execute('SELECT SUM(cantidad_vendida) FROM productos_sj')
    total_ventas = cursor.fetchone()[0]

    # Calcular el porcentaje de ventas para cada producto
    productos_con_porcentaje = []
    for producto in productos:
        nombre, precio, cantidad_vendida = producto
        porcentaje = (cantidad_vendida / total_ventas) * 100 if total_ventas > 0 else 0
        productos_con_porcentaje.append({
            'nombre': nombre,
            'precio': precio,
            'cantidad_vendida': cantidad_vendida,
            'porcentaje': round(porcentaje, 2)  # Redondear a 2 decimales
        })

    # Cerrar la conexión
    conn.close()

    # Renderizar la plantilla HTML con los productos y el total de ventas
    return render_template('productos_mas_vendidos.html', productos=productos_con_porcentaje, total_ventas=total_ventas)

# Ruta para productos por agotarse
@app.route('/productos_por_agotarse')
def productos_por_agotarse():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Obtener productos con stock menor o igual a 2
    cursor.execute('''
    SELECT id, nombre, codigo_barras, stock, precio, precio_costo
    FROM productos_sj
    WHERE stock <= 2
    ORDER BY stock ASC
    ''')
    productos = cursor.fetchall()

    conn.close()
    return render_template('productos_por_agotarse.html', productos=productos)

# Ruta principal para mostrar las ventas y reparaciones
from flask import send_file
from openpyxl import Workbook
from io import BytesIO

@app.route('/ultimas_ventas')
def ultimas_ventas():
    conn = get_db_connection()
    cursor = conn.cursor()

    argentina_tz = pytz.timezone('America/Argentina/Buenos_Aires')
    fecha_actual = datetime.now(argentina_tz).strftime('%Y-%m-%d')

    fecha_desde = request.args.get('fecha_desde', fecha_actual)
    fecha_hasta = request.args.get('fecha_hasta', fecha_actual)
    exportar = str(request.args.get('exportar', '')).lower() in ('1', 'true', 'on', 'sí', 'si')

    # Ventas: calcular el precio real según tipo_precio
    cursor.execute('''
        SELECT 
            v.id AS venta_id,
            p.nombre AS nombre_producto,
            p.num AS num,  -- <<--- agregado
            v.cantidad,
            CASE
                WHEN v.tipo_precio = 'revendedor' THEN p.precio_revendedor
                ELSE p.precio
            END AS precio_unitario,
            v.cantidad * 
            CASE
                WHEN v.tipo_precio = 'revendedor' THEN p.precio_revendedor
                ELSE p.precio
            END AS total,
            v.fecha,
            v.tipo_pago,
            v.dni_cliente,
            v.tipo_precio
        FROM ventas_sj v
        LEFT JOIN productos_sj p ON v.producto_id = p.id
        WHERE DATE(v.fecha) BETWEEN %s AND %s
        ORDER BY v.fecha DESC
    ''', (fecha_desde, fecha_hasta))
    ventas = cursor.fetchall()

    # Reparaciones
    cursor.execute('''
        SELECT 
            id AS reparacion_id,
            nombre_servicio,
            cantidad,
            precio AS precio_unitario,
            (cantidad * precio) AS total,
            fecha,
            tipo_pago
        FROM reparaciones_sj
        WHERE DATE(fecha) BETWEEN %s AND %s
        ORDER BY fecha DESC
    ''', (fecha_desde, fecha_hasta))
    reparaciones = cursor.fetchall()

    # Totales por tipo de pago
    total_ventas_por_pago = {}
    for venta in ventas:
        monto = venta['total'] or 0
        total_ventas_por_pago[venta['tipo_pago']] = total_ventas_por_pago.get(venta['tipo_pago'], 0) + monto

    total_reparaciones_por_pago = {}
    for reparacion in reparaciones:
        total_reparaciones_por_pago[reparacion['tipo_pago']] = total_reparaciones_por_pago.get(reparacion['tipo_pago'], 0) + (reparacion['total'] or 0)

    if exportar:
        from openpyxl import Workbook
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"

        headers = [
            "ID Venta", "Producto", "Cantidad", "Núm",  # <<--- agregado "Núm"
            "Precio Unitario", "Total", "Fecha", "Tipo de Pago",
            "DNI Cliente", "Tipo Precio"
        ]
        ws.append(headers)

        for venta in ventas:
            ws.append([
                venta['venta_id'],
                venta['nombre_producto'],
                venta['cantidad'],
                venta.get('num') or '',  # <<--- agregado
                venta['precio_unitario'] or 0,
                venta['total'] or 0,
                venta['fecha'].strftime('%Y-%m-%d %H:%M:%S') if venta['fecha'] else '',
                venta['tipo_pago'],
                venta['dni_cliente'] or '',
                venta['tipo_precio'].capitalize() if venta['tipo_precio'] else ''
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        nombre_archivo = f"ventas_{fecha_desde}_a_{fecha_hasta}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    conn.close()
    return render_template(
        'ultimas_ventas.html',
        ventas=ventas,
        reparaciones=reparaciones,
        fecha_actual=fecha_actual,
        total_ventas_por_pago=total_ventas_por_pago,
        total_reparaciones_por_pago=total_reparaciones_por_pago,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta
    )



# Ruta para eliminar una venta
@app.route('/anular_venta/<int:venta_id>', methods=['POST'])
def anular_venta(venta_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Verificar si la venta existe
        cursor.execute('SELECT * FROM ventas_sj WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()

        if not venta:
            return jsonify({'success': False, 'message': 'Venta no encontrada'}), 404

        # Eliminar la venta
        cursor.execute('DELETE FROM ventas_sj WHERE id = %s', (venta_id,))
        conn.commit()

        return jsonify({'success': True, 'message': 'Venta eliminada correctamente'}), 200
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

# Ruta para eliminar una reparación
@app.route('/anular_reparacion/<int:reparacion_id>', methods=['POST'])
def anular_reparacion(reparacion_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Verificar si la reparación existe
        cursor.execute('SELECT * FROM reparaciones_sj WHERE id = %s', (reparacion_id,))
        reparacion = cursor.fetchone()

        if not reparacion:
            return jsonify({'success': False, 'message': 'Reparación no encontrada'}), 404

        # Eliminar la reparación
        cursor.execute('DELETE FROM reparaciones_sj WHERE id = %s', (reparacion_id,))
        conn.commit()

        return jsonify({'success': True, 'message': 'Reparación eliminada correctamente'}), 200
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

# Ruta para egresos
@app.route('/egresos', methods=['GET', 'POST'])
def egresos():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Agregar un nuevo egreso
    if request.method == 'POST' and 'agregar' in request.form:
        fecha = request.form['fecha']
        monto = float(request.form['monto'])
        descripcion = request.form['descripcion']
        tipo_pago = request.form['tipo_pago']  # Nuevo campo

        cursor.execute('''
        INSERT INTO egresos_sj (fecha, monto, descripcion, tipo_pago)
        VALUES (%s, %s, %s, %s)
        ''', (fecha, monto, descripcion, tipo_pago))
        conn.commit()
        conn.close()
        return redirect(url_for('egresos'))

    # Eliminar un egreso
    if request.method == 'POST' and 'eliminar' in request.form:
        egreso_id = request.form['egreso_id']
        cursor.execute('DELETE FROM egresos_sj WHERE id = %s', (egreso_id,))
        conn.commit()
        conn.close()
        return redirect(url_for('egresos'))

    # Obtener todos los egresos
    cursor.execute('SELECT id, fecha, monto, descripcion, tipo_pago FROM egresos_sj ORDER BY fecha DESC')
    egresos = cursor.fetchall()
    conn.close()
    return render_template('egresos.html', egresos=egresos)

@app.route('/dashboard')
def dashboard():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Obtener fechas
    fecha_desde = request.args.get('fecha_desde', datetime.now().strftime('%Y-%m-%d'))
    fecha_hasta = request.args.get('fecha_hasta', datetime.now().strftime('%Y-%m-%d'))

    # Total de ventas de productos según tipo_precio
    cursor.execute('''
    SELECT SUM(
        v.cantidad *
        CASE
            WHEN v.tipo_precio = 'revendedor' THEN p.precio_revendedor
            ELSE p.precio
        END
    ) AS total_ventas_productos
    FROM ventas_sj v
    LEFT JOIN productos_sj p ON v.producto_id = p.id
    WHERE DATE(v.fecha) BETWEEN %s AND %s
    ''', (fecha_desde, fecha_hasta))
    total_ventas_productos = cursor.fetchone()['total_ventas_productos'] or 0

    # Total de ventas de reparaciones
    cursor.execute('''
    SELECT SUM(precio) AS total_ventas_reparaciones
    FROM reparaciones_sj
    WHERE DATE(fecha) BETWEEN %s AND %s
    ''', (fecha_desde, fecha_hasta))
    total_ventas_reparaciones = cursor.fetchone()['total_ventas_reparaciones'] or 0

    total_ventas = total_ventas_productos + total_ventas_reparaciones

    # Total de egresos
    cursor.execute('''
    SELECT SUM(monto) AS total_egresos
    FROM egresos_sj
    WHERE DATE(fecha) BETWEEN %s AND %s
    ''', (fecha_desde, fecha_hasta))
    total_egresos = cursor.fetchone()['total_egresos'] or 0

    # Costo total de productos vendidos
    cursor.execute('''
    SELECT SUM(v.cantidad * p.precio_costo) AS total_costo
    FROM ventas_sj v
    JOIN productos_sj p ON v.producto_id = p.id
    WHERE DATE(v.fecha) BETWEEN %s AND %s
    ''', (fecha_desde, fecha_hasta))
    total_costo = cursor.fetchone()['total_costo'] or 0

    # Ganancia
    ganancia = total_ventas - total_egresos - total_costo

    # Distribución de ventas
    cursor.execute('''
    SELECT 'Productos' AS tipo, SUM(
        v.cantidad *
        CASE
            WHEN v.tipo_precio = 'revendedor' THEN p.precio_revendedor
            ELSE p.precio
        END
    ) AS total
    FROM ventas_sj v
    LEFT JOIN productos_sj p ON v.producto_id = p.id
    WHERE DATE(v.fecha) BETWEEN %s AND %s
    UNION ALL
    SELECT 'Reparaciones' AS tipo, SUM(precio)
    FROM reparaciones_sj
    WHERE DATE(fecha) BETWEEN %s AND %s
    ''', (fecha_desde, fecha_hasta, fecha_desde, fecha_hasta))
    distribucion_ventas = cursor.fetchall()

    conn.close()

    return render_template('dashboard.html', 
                          total_ventas=total_ventas, 
                          total_egresos=total_egresos, 
                          total_costo=total_costo, 
                          ganancia=ganancia,
                          total_ventas_productos=total_ventas_productos,
                          total_ventas_reparaciones=total_ventas_reparaciones,
                          distribucion_ventas=distribucion_ventas,
                          fecha_desde=fecha_desde,
                          fecha_hasta=fecha_hasta)

# Ruta para resumen semanal
@app.route('/resumen_semanal')
def resumen_semanal():
    # Obtener la fecha de inicio de la semana (lunes)
    hoy = datetime.now()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    inicio_semana_str = inicio_semana.strftime('%Y-%m-%d')

    # Conectar a la base de datos
    conn = get_db_connection()
    cursor = conn.cursor()

    # Consultar las ventas de la semana actual
    cursor.execute('''
        SELECT tipo_pago, SUM(total) as total
        FROM ventas_sj
        WHERE fecha >= %s
        GROUP BY tipo_pago
    ''', (inicio_semana_str,))

    resumen = cursor.fetchall()

    # Cerrar la conexión
    conn.close()

    # Renderizar la plantilla con el resumen
    return render_template('resumen_semanal.html', resumen=resumen)

@app.route('/caja')
def caja():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fecha actual Argentina
    argentina_tz = pytz.timezone('America/Argentina/Buenos_Aires')
    hoy = datetime.now(argentina_tz).date()

    # Fechas desde y hasta
    fecha_desde = request.args.get('fecha_desde', hoy.strftime('%Y-%m-%d'))
    fecha_hasta = request.args.get('fecha_hasta', hoy.strftime('%Y-%m-%d'))

    # Ventas: calcular total según tipo_precio
    cursor.execute('''
        SELECT 
            v.id AS venta_id,
            p.nombre AS nombre_producto,
            v.cantidad,
            v.tipo_precio,
            v.tipo_pago,
            v.fecha,
            CASE
                WHEN v.tipo_precio = 'revendedor' THEN p.precio_revendedor
                ELSE p.precio
            END AS precio_unitario,
            (v.cantidad * 
             CASE
                WHEN v.tipo_precio = 'revendedor' THEN p.precio_revendedor
                ELSE p.precio
             END) AS total
        FROM ventas_sj v
        JOIN productos_sj p ON v.producto_id = p.id
        WHERE DATE(v.fecha) BETWEEN %s AND %s
        ORDER BY v.fecha DESC
    ''', (fecha_desde, fecha_hasta))
    ventas = cursor.fetchall()

    # Reparaciones
    cursor.execute('''
        SELECT 
            id AS reparacion_id,
            nombre_servicio AS nombre_servicio,
            cantidad,
            precio AS precio_unitario,
            (cantidad * precio) AS total,
            fecha,
            tipo_pago
        FROM reparaciones_sj
        WHERE DATE(fecha) BETWEEN %s AND %s
        ORDER BY fecha DESC
    ''', (fecha_desde, fecha_hasta))
    reparaciones = cursor.fetchall()

    # Egresos
    cursor.execute('''
        SELECT 
            id AS egreso_id,
            descripcion,
            monto,
            tipo_pago,
            fecha
        FROM egresos_sj
        WHERE DATE(fecha) BETWEEN %s AND %s
        ORDER BY fecha DESC
    ''', (fecha_desde, fecha_hasta))
    egresos = cursor.fetchall()

    # Totales por tipo de pago
    total_ventas_por_pago = {}
    for venta in ventas:
        tipo_pago = venta['tipo_pago']
        total = venta['total']
        total_ventas_por_pago[tipo_pago] = total_ventas_por_pago.get(tipo_pago, 0) + total

    total_reparaciones_por_pago = {}
    for r in reparaciones:
        tipo_pago = r['tipo_pago']
        total = r['total']
        total_reparaciones_por_pago[tipo_pago] = total_reparaciones_por_pago.get(tipo_pago, 0) + total

    total_combinado_por_pago = total_ventas_por_pago.copy()
    for tipo_pago, total in total_reparaciones_por_pago.items():
        total_combinado_por_pago[tipo_pago] = total_combinado_por_pago.get(tipo_pago, 0) + total

    total_egresos_por_pago = {}
    for egreso in egresos:
        tipo_pago = egreso['tipo_pago']
        monto = egreso['monto']
        total_egresos_por_pago[tipo_pago] = total_egresos_por_pago.get(tipo_pago, 0) + monto

    # Neto
    neto_por_pago = {}
    for tipo_pago, total in total_combinado_por_pago.items():
        egresos = total_egresos_por_pago.get(tipo_pago, 0)
        neto_por_pago[tipo_pago] = total - egresos

    conn.close()

    return render_template(
        'caja.html',
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        neto_por_pago=neto_por_pago
    )


# Ruta para reparaciones
import unicodedata

def normalizar(texto):
    return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode().lower().strip()

@app.route('/reparaciones', methods=['GET', 'POST'])
def reparaciones():
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'POST':
        tipo_reparacion = request.form['tipo_reparacion']
        marca = request.form['equipo']
        modelo = request.form['modelo']
        tecnico = request.form['tecnico']
        monto = float(request.form['monto'])
        nombre_cliente = request.form['nombre_cliente']
        telefono = request.form['telefono']
        nro_orden = request.form['nro_orden']
        fecha = datetime.now().date()
        hora = datetime.now().strftime('%H:%M:%S')
        estado = 'Por Reparar'

        cursor.execute('''
            INSERT INTO equipos_sj (
                tipo_reparacion, marca, modelo, tecnico, monto,
                nombre_cliente, telefono, nro_orden, fecha, hora, estado
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            tipo_reparacion, marca, modelo, tecnico, monto,
            nombre_cliente, telefono, nro_orden, fecha, hora, estado
        ))
        conn.commit()

    # Fechas desde GET
    fecha_desde = request.args.get('fecha_desde', (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
    fecha_hasta = request.args.get('fecha_hasta', datetime.now().strftime('%Y-%m-%d'))

    # Últimos equipos
    cursor.execute('''
        SELECT * FROM equipos_sj
        WHERE fecha BETWEEN %s AND %s
        ORDER BY nro_orden DESC
    ''', (fecha_desde, fecha_hasta))
    ultimos_equipos = cursor.fetchall()

    # Por técnico
    cursor.execute('''
        SELECT tecnico, COUNT(*) as cantidad
        FROM equipos_sj
        WHERE fecha BETWEEN %s AND %s
        GROUP BY tecnico
    ''', (fecha_desde, fecha_hasta))
    datos_tecnicos = cursor.fetchall()
    equipos_por_tecnico = {row['tecnico']: row['cantidad'] for row in datos_tecnicos}

    # Por estado
    cursor.execute('''
        SELECT estado, COUNT(*) as cantidad
        FROM equipos_sj
        WHERE fecha BETWEEN %s AND %s
        GROUP BY estado
    ''', (fecha_desde, fecha_hasta))
    datos_estados = cursor.fetchall()

    # Inicializar resumen
    estados = {
        'por_reparar': 0,
        'en_reparacion': 0,
        'listo': 0,
        'retirado': 0,
        'no_salio': 0
    }

    for row in datos_estados:
        estado = normalizar(row['estado'])
        cantidad = row['cantidad']

        if estado in ['por reparar', 'por_reparar']:
            estados['por_reparar'] += cantidad
        elif estado in ['en reparacion', 'en reparación', 'en_reparacion']:
            estados['en_reparacion'] += cantidad
        elif estado == 'listo':
            estados['listo'] += cantidad
        elif estado == 'retirado':
            estados['retirado'] += cantidad
        elif estado in ['no salio', 'no_salio']:
            estados['no_salio'] += cantidad
        else:
            # Por si aparece un nuevo estado no previsto
            estados[estado] = cantidad

    # Total sumando todos menos la clave 'total'
    estados['total'] = sum([
        cantidad for key, cantidad in estados.items()
        if key != 'total'
    ])

    conn.close()

    return render_template(
        'reparaciones.html',
        ultimos_equipos=ultimos_equipos,
        equipos_por_tecnico=equipos_por_tecnico,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        estados=estados
    )



# Ruta para eliminar reparaciones
@app.route('/eliminar_reparacion/<int:id>', methods=['POST'])
def eliminar_reparacion(id):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Eliminar el equipo por su ID
    cursor.execute('DELETE FROM equipos_sj WHERE id = %s', (id,))
    conn.commit()
    conn.close()

    # Redirigir a la página de reparaciones después de eliminar
    return redirect(url_for('reparaciones'))

# Ruta para actualizar estado de reparaciones
@app.route('/actualizar_estado', methods=['POST'])
def actualizar_estado():
    data = request.get_json()
    nro_orden = data['nro_orden']
    estado = data['estado']

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE equipos_sj
        SET estado = %s
        WHERE nro_orden = %s
    ''', (estado, nro_orden))
    conn.commit()
    conn.close()

    return jsonify({'success': True})

# Ruta para mercadería fallada
@app.route('/mercaderia_fallada', methods=['GET', 'POST'])
def mercaderia_fallada():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Buscar productos
    if request.method == 'POST' and 'buscar' in request.form:
        busqueda = request.form['busqueda']
        cursor.execute('''
        SELECT id, nombre, codigo_barras, stock, precio, precio_costo
        FROM productos_sj
        WHERE nombre LIKE %s OR codigo_barras LIKE %s
        ''', (f'%{busqueda}%', f'%{busqueda}%'))
        productos = cursor.fetchall()
        conn.close()
        return render_template('mercaderia_fallada.html', productos=productos)

    # Registrar mercadería fallada
    if request.method == 'POST' and 'registrar_fallada' in request.form:
        producto_id = request.form['producto_id']
        cantidad = int(request.form['cantidad'])
        descripcion = request.form['descripcion']
        fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Verificar si hay suficiente stock
        cursor.execute('SELECT stock FROM productos_sj WHERE id = %s', (producto_id,))
        producto = cursor.fetchone()

        if producto and producto['stock'] >= cantidad:
            # Registrar en la tabla `mercaderia_fallada`
            cursor.execute('''
            INSERT INTO mercaderia_fallada (producto_id, cantidad, fecha, descripcion)
            VALUES (%s, %s, %s, %s)
            ''', (producto_id, cantidad, fecha, descripcion))

            # Actualizar el stock en la tabla `productos`
            cursor.execute('UPDATE productos_sj SET stock = stock - %s WHERE id = %s', (cantidad, producto_id))
            conn.commit()
            conn.close()
            return redirect(url_for('mercaderia_fallada'))
        else:
            conn.close()
            return f"No hay suficiente stock para el producto seleccionado."

    # Obtener historial de mercadería fallada
    cursor.execute('''
    SELECT mf.id, p.nombre, mf.cantidad, mf.fecha, mf.descripcion
    FROM mercaderia_fallada mf
    JOIN productos_sj p ON mf.producto_id = p.id
    ORDER BY mf.fecha DESC
    ''')
    historial = cursor.fetchall()

    conn.close()
    return render_template('mercaderia_fallada.html', historial=historial)





@app.route('/agregar_stock', methods=['GET', 'POST'])
def agregar_stock():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute('SELECT nombre FROM categorias_sj ORDER BY nombre')
    categorias = [row['nombre'] for row in cursor.fetchall()]

    # Agregar nueva categoría
    if request.method == 'POST' and 'nueva_categoria' in request.form:
        nueva_categoria = request.form['nueva_categoria'].strip()
        if nueva_categoria:
            cursor.execute('INSERT INTO categorias_sj (nombre) VALUES (%s) ON CONFLICT DO NOTHING', (nueva_categoria,))
            conn.commit()
        return redirect(url_for('agregar_stock'))

    # Eliminar categoría
    if request.method == 'POST' and 'eliminar_categoria' in request.form:
        categoria_a_eliminar = request.form['eliminar_categoria']
        cursor.execute('DELETE FROM categorias_sj WHERE nombre = %s', (categoria_a_eliminar,))
        conn.commit()
        return redirect(url_for('agregar_stock'))

    # Obtener el término de búsqueda (si existe)
    busqueda = request.args.get('busqueda', '')

    try:
        # Eliminar un producto
        if request.method == 'POST' and 'eliminar' in request.form:
            producto_id = request.form['producto_id']
            cursor.execute('DELETE FROM productos_sj WHERE id = %s', (producto_id,))
            conn.commit()
            return redirect(url_for('agregar_stock'))

        # Editar un producto
        if request.method == 'POST' and 'editar' in request.form:
            producto_id = request.form['producto_id']
            nombre = request.form['nombre'].upper()
            codigo_barras = request.form['codigo_barras']
            stock = int(request.form['stock'])
            precio = float(request.form['precio'])
            precio_costo = float(request.form['precio_costo'])
            categoria = request.form.get('categoria')
            num = request.form.get('num')
            color = request.form.get('color')
            bateria = request.form.get('bateria')
            precio_revendedor = request.form.get('precio_revendedor')
            condicion = request.form.get('condicion')

            if precio_revendedor:
                try:
                    precio_revendedor = float(precio_revendedor)
                except:
                    precio_revendedor = None
            else:
                precio_revendedor = None

            foto_url = None
            if 'foto' in request.files:
                foto = request.files['foto']
                if foto and foto.filename != '':
                    result = cloudinary.uploader.upload(foto)
                    foto_url = result['secure_url']

            if foto_url:
                cursor.execute('''
                    UPDATE productos_sj
                    SET nombre=%s, codigo_barras=%s, stock=%s, precio=%s, precio_costo=%s,
                        categoria=%s, foto_url=%s, num=%s, color=%s, bateria=%s, precio_revendedor=%s, condicion=%s
                    WHERE id=%s
                ''', (nombre, codigo_barras, stock, precio, precio_costo,
                      categoria, foto_url, num, color, bateria, precio_revendedor, condicion, producto_id))
            else:
                cursor.execute('''
                    UPDATE productos_sj
                    SET nombre=%s, codigo_barras=%s, stock=%s, precio=%s, precio_costo=%s,
                        categoria=%s, num=%s, color=%s, bateria=%s, precio_revendedor=%s, condicion=%s
                    WHERE id=%s
                ''', (nombre, codigo_barras, stock, precio, precio_costo,
                      categoria, num, color, bateria, precio_revendedor, condicion, producto_id))

            conn.commit()
            return redirect(url_for('agregar_stock'))

        # Agregar stock a un producto existente
        if request.method == 'POST' and 'agregar_stock' in request.form:
            producto_id = request.form['producto_id']
            cantidad = int(request.form['cantidad'])

            cursor.execute('''
                UPDATE productos_sj
                SET stock = stock + %s
                WHERE id = %s
            ''', (cantidad, producto_id))
            conn.commit()
            return redirect(url_for('agregar_stock'))

        # Agregar un nuevo producto
        if request.method == 'POST' and 'agregar' in request.form:
            nombre = request.form['nombre'].upper()
            codigo_barras = request.form['codigo_barras']
            stock = int(request.form['stock'])
            precio = float(request.form['precio'])
            precio_costo = float(request.form['precio_costo'])
            categoria = request.form.get('categoria')
            num = request.form.get('num')
            color = request.form.get('color')
            bateria = request.form.get('bateria')
            precio_revendedor = request.form.get('precio_revendedor')
            condicion = request.form.get('condicion')

            if precio_revendedor:
                try:
                    precio_revendedor = float(precio_revendedor)
                except:
                    precio_revendedor = None
            else:
                precio_revendedor = None

            foto_url = None
            if 'foto' in request.files:
                foto = request.files['foto']
                if foto and foto.filename != '':
                    result = cloudinary.uploader.upload(foto)
                    foto_url = result['secure_url']

            cursor.execute('''
                INSERT INTO productos_sj (
                    nombre, codigo_barras, stock, precio, precio_costo,
                    foto_url, categoria, num, color, bateria, precio_revendedor, condicion
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (nombre, codigo_barras, stock, precio, precio_costo,
                  foto_url, categoria, num, color, bateria, precio_revendedor, condicion))
            conn.commit()
            return redirect(url_for('agregar_stock'))

        # Obtener productos
        if busqueda:
            cursor.execute('''
                SELECT id, nombre, codigo_barras, stock, precio, precio_costo, foto_url,
                       num, color, bateria, precio_revendedor, condicion
                FROM productos_sj
                WHERE nombre ILIKE %s OR codigo_barras ILIKE %s
            ''', (f'%{busqueda}%', f'%{busqueda}%'))
        else:
            cursor.execute('''
                SELECT id, nombre, codigo_barras, stock, precio, precio_costo, foto_url,
                       num, color, bateria, precio_revendedor, condicion
                FROM productos_sj
            ''')

        productos = cursor.fetchall()
        return render_template('agregar_stock.html', productos=productos, busqueda=busqueda, categorias=categorias)

    except Exception as e:
        conn.rollback()
        return f"Error: {str(e)}"
    finally:
        conn.close()




@app.route('/tienda')
def tienda():
    conn = get_db_connection()
    cursor = conn.cursor()

    categoria = request.args.get('categoria')

    if categoria:
        cursor.execute('''
            SELECT id, nombre, codigo_barras, stock, precio, foto_url, categoria,
                   color, bateria, condicion, precio_revendedor
            FROM productos_sj
            WHERE categoria = %s AND foto_url IS NOT NULL AND stock > 0
            ORDER BY nombre
        ''', (categoria,))
    else:
        cursor.execute('''
            SELECT id, nombre, codigo_barras, stock, precio, foto_url, categoria,
                   color, bateria, condicion, precio_revendedor
            FROM productos_sj
            WHERE foto_url IS NOT NULL AND stock > 0
            ORDER BY nombre
        ''')

    productos = cursor.fetchall()

    cursor.execute('SELECT DISTINCT categoria FROM productos_sj WHERE categoria IS NOT NULL ORDER BY categoria')
    categorias = [row[0] for row in cursor.fetchall()]

    conn.close()
    return render_template(
        'tienda.html',
        productos=productos,
        categorias=categorias,
        categoria_seleccionada=categoria
    )


def crear_tabla_categorias():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS categorias_sj (
            id SERIAL PRIMARY KEY,
            nombre TEXT UNIQUE NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

crear_tabla_categorias()


from openpyxl.drawing.image import Image as XLImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from flask import send_file
from io import BytesIO

@app.route('/exportar_stock')
def exportar_stock():
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl import Workbook
    from io import BytesIO

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, nombre, codigo_barras, num, color, bateria, condicion, stock, precio, precio_costo, precio_revendedor
        FROM productos_sj
        ORDER BY nombre
    """)
    productos = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    # Encabezados
    headers = [
        "ID", "Nombre", "Código", "Núm", "Color", "Batería",
        "Condición", "Stock", "Precio Venta", "Precio Costo", "Precio Rev."
    ]
    ws.append(headers)

    # Estilo de encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar filas
    for fila in productos:
        ws.append(fila)

    # Ajustar ancho automático
    for column in ws.columns:
        max_len = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws.column_dimensions[column[0].column_letter].width = max_len + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="stock.xlsx", as_attachment=True)





if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)

#if __name__ == '__main__': para uso local
 #   app.run(debug=True)
